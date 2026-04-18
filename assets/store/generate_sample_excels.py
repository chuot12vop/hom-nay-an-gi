from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent
HEADER_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
HEADER_FONT = Font(bold=True)


def write_xlsx(filename: str, headers: list[str], rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = filename.replace(".xlsx", "")

    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for col in sheet.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        sheet.column_dimensions[col[0].column_letter].width = min(max(12, max_len + 2), 40)

    workbook.save(BASE_DIR / filename)


def main() -> None:
    write_xlsx(
        "meals.xlsx",
        ["meal_id", "meal_code", "meal_name_vi", "meal_name_en", "description"],
        [
            [1, "BREAKFAST", "Bua sang", "Breakfast", "Bua an dau ngay"],
            [2, "LUNCH", "Bua trua", "Lunch", "Bua chinh buoi trua"],
            [3, "DINNER", "Bua toi", "Dinner", "Bua chinh buoi toi"],
            [4, "SNACK", "Bua phu", "Snack", "Bua an nhe"],
        ],
    )

    write_xlsx(
        "food_categories.xlsx",
        ["category_id", "category_code", "category_name_vi", "category_name_en", "is_active"],
        [
            [1, "NOODLE", "Mon nuoc", "Noodle Soup", True],
            [2, "RICE", "Mon com", "Rice Dishes", True],
            [3, "GRILL", "Mon nuong", "Grilled Dishes", True],
            [4, "VEGAN", "Mon chay", "Vegan Dishes", True],
            [5, "FAST", "Do an nhanh", "Fast Food", True],
        ],
    )

    write_xlsx(
        "ingredients.xlsx",
        [
            "ingredient_id",
            "ingredient_name_vi",
            "unit",
            "default_quantity",
            "calories_per_unit",
            "is_vegetarian",
        ],
        [
            [1, "Gao", "gram", 100, 130, True],
            [2, "Thit ga", "gram", 100, 165, False],
            [3, "Thit bo", "gram", 100, 250, False],
            [4, "Tom", "gram", 100, 99, False],
            [5, "Dau hu", "gram", 100, 76, True],
            [6, "Rau cai", "gram", 100, 25, True],
            [7, "Mi", "gram", 100, 138, True],
            [8, "Trung ga", "piece", 1, 72, False],
        ],
    )

    write_xlsx(
        "foods.xlsx",
        [
            "food_id",
            "food_name_vi",
            "food_name_en",
            "category_id",
            "meal_id",
            "main_ingredient_id",
            "price_vnd",
            "spicy_level",
            "is_active",
        ],
        [
            [1, "Pho bo", "Beef Pho", 1, 2, 3, 50000, 1, True],
            [2, "Bun cha", "Grilled Pork Vermicelli", 1, 2, 2, 45000, 2, True],
            [3, "Com tam suon", "Broken Rice with Pork", 2, 2, 2, 55000, 1, True],
            [4, "Com ga xoi mo", "Fried Chicken Rice", 2, 3, 2, 60000, 1, True],
            [5, "Dau hu sot ca", "Tofu in Tomato Sauce", 4, 3, 5, 40000, 0, True],
            [6, "Mi xao rau", "Stir-fried Noodles", 1, 4, 7, 35000, 1, True],
            [7, "Hamburger bo", "Beef Burger", 5, 4, 3, 65000, 1, True],
            [8, "Salad tom", "Shrimp Salad", 4, 1, 4, 52000, 0, True],
        ],
    )

    print("Created sample Excel files in assets/store")


if __name__ == "__main__":
    main()
